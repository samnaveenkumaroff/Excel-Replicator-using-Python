import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
from openpyxl import load_workbook
#operate
class FileOperations:
    def __init__(self):
        pass

    def load_sheet_names(self, file_path, combobox):
        try:
            workbook = load_workbook(file_path)
            sheets = workbook.sheetnames
            combobox['values'] = sheets
            if sheets:
                combobox.current(0)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheets: {str(e)}")

    def view_file(self, file_path, combobox):
        try:
            if os.path.exists(file_path):
                workbook = load_workbook(file_path, read_only=True)
                sheets = workbook.sheetnames
                combobox['values'] = sheets
                if sheets:
                    combobox.current(0)
            else:
                messagebox.showerror("Error", "File not found.")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
