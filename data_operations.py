import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
from openpyxl import load_workbook

class DataOperations:
    def __init__(self):
        pass

    def pull_data(self, source_file, target_file, entire_file_var, source_sheet_combobox, target_sheet_combobox):
        try:
            source_workbook = load_workbook(source_file, read_only=True)
            target_workbook = load_workbook(target_file)

            if entire_file_var:
                # Copy entire workbook from source to target
                for source_sheet in source_workbook.sheetnames:
                    if source_sheet not in target_workbook.sheetnames:
                        target_workbook.create_sheet(title=source_sheet)

                    source_sheet_data = source_workbook[source_sheet]
                    target_sheet_data = target_workbook[source_sheet]

                    for row in source_sheet_data.iter_rows(values_only=True):
                        target_sheet_data.append(row)

            else:
                # Copy selected sheet from source to target
                source_sheet = source_sheet_combobox.get()
                target_sheet = target_sheet_combobox.get()

                if not source_sheet or not target_sheet:
                    messagebox.showerror("Error", "Please select source and target sheets.")
                    return

                if source_sheet not in source_workbook.sheetnames:
                    messagebox.showerror("Error", f"Sheet '{source_sheet}' does not exist in '{os.path.basename(source_file)}'.")
                    return

                if target_sheet not in target_workbook.sheetnames:
                    messagebox.showerror("Error", f"Sheet '{target_sheet}' does not exist in '{os.path.basename(target_file)}'.")
                    return

                source_sheet_data = source_workbook[source_sheet]
                target_sheet_data = target_workbook[target_sheet]

                for row in source_sheet_data.iter_rows(values_only=True):
                    target_sheet_data.append(row)

            target_workbook.save(target_file)
            messagebox.showinfo("Success", "Data pulled successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to pull data: {str(e)}")

    def replicate_data(self, source_file, target_file, entire_file_var, source_sheet_combobox, target_sheet_combobox):
        try:
            source_workbook = load_workbook(source_file, read_only=True)
            target_workbook = load_workbook(target_file)

            if entire_file_var:
                # Copy entire workbook from source to target
                for source_sheet in source_workbook.sheetnames:
                    if source_sheet not in target_workbook.sheetnames:
                        target_workbook.create_sheet(title=source_sheet)

                    source_sheet_data = source_workbook[source_sheet]
                    target_sheet_data = target_workbook[source_sheet]

                    for row in source_sheet_data.iter_rows(values_only=True):
                        target_sheet_data.append(row)

            else:
                # Copy selected sheet from source to target
                source_sheet = source_sheet_combobox.get()
                target_sheet = target_sheet_combobox.get()

                if not source_sheet or not target_sheet:
                    messagebox.showerror("Error", "Please select source and target sheets.")
                    return

                if source_sheet not in source_workbook.sheetnames:
                    messagebox.showerror("Error", f"Sheet '{source_sheet}' does not exist in '{os.path.basename(source_file)}'.")
                    return

                if target_sheet not in target_workbook.sheetnames:
                    messagebox.showerror("Error", f"Sheet '{target_sheet}' does not exist in '{os.path.basename(target_file)}'.")
                    return

                source_sheet_data = source_workbook[source_sheet]
                target_sheet_data = target_workbook[target_sheet]

                for row in source_sheet_data.iter_rows(values_only=True):
                    target_sheet_data.append(row)

            target_workbook.save(target_file)
            messagebox.showinfo("Success", "Data replicated successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to replicate data: {str(e)}")
